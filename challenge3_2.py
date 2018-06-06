from openpyxl import load_workbook
from openpyxl import Workbook

from datetime import datetime

wb = load_workbook('courses.xlsx')
student_sheet = wb['students']
time_sheet = wb['time']


def combine():
    combine_sheet = wb.create_sheet(title='combine')
    combine_sheet.append(['创建时间', '课程名称', '学习人数', '学习时间'])

    for stu in student_sheet.values:
        if stu[2] != '学习人数':
            for time in time_sheet.values:
                if time[1] == stu[2]:
                    combine_sheet.append(list(shu) + time[2])

    wb.save('courses.xlsx')

def split():
    combine_sheet = wb['combine']
    split_name = []

    for item in combine_sheet.values:
        if item[0] != '创建时间':
            split_name.append(item[0].strftime('%Y'))

    for name in set(split_name):
        wb_temp = Workbook()
        wb_temp.remove(wb_temp.active)

        ws = wb_temp.create_sheet(title=name)

        for item_by_year in combine_sheet_year:
            if item_by_year[0] != '创建时间':
                if item_by_year[0].strftime('%Y') == name:
                    ws.append(item_by_year)
        
        wb_temp.save('{}.xlsx'.format(name))


if __name__ == '__main__':
    combine() 
    split()
